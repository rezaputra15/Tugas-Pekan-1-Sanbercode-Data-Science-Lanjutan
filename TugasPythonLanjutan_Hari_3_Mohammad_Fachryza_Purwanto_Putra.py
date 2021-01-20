from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
import csv, pandas as pd

wb = Workbook(write_only=True)
ws = wb.create_sheet()

data1 = pd.read_csv(r"jumlah-penduduk-kota-bandung.csv")
data2 = pd.read_csv(r"luas-wilayah-menurut-kecamatan-di-kota-bandung-2017.csv")

df1 = data1.sort_values('Kecamatan  ', ignore_index=True)
df2 = data2.sort_values('Nama Kecamatan', ignore_index=True)
df2 = df2['Luas Wilayah (m2)'] / 100
new_df2 = pd.to_numeric(df2, downcast='signed')
df = pd.merge(df1, new_df2, left_index=True, right_index=True)

df['Kepadatan Penduduk'] = df['Jumlah_Penduduk'] / df['Luas Wilayah (m2)']
df = df[['Kecamatan  ', 'Kepadatan Penduduk']]

df.to_csv(r"kepadatan-penduduk-bandung.csv", index=False)

data = open(r"kepadatan-penduduk-bandung.csv")
rows = csv.reader(data, delimiter=',')

for row in rows:
    data_clean = []
    for i in row:
        try:
            i = float(i)
        except:
            pass
        data_clean.append(i)
    ws.append(data_clean)

chart1 = BarChart()
chart1.type = "col"
chart1.style = 5
chart1.title = "Tingkat Kepadatan Penduduk Bandung Tahun 2017"
chart1.y_axis.title = "Kepadatan Penduduk"
chart1.x_axis.title = "Kecamatan  "

data = Reference(ws, min_col=2, min_row=1, max_row=31)
cats = Reference(ws, min_col=1, min_row=2, max_row=31)

chart1.height = 10
chart1.width = 30
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
ws.add_chart(chart1, "E2")

wb.save(r"Fachryza-chart.xlsx")